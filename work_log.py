import os
import datetime
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, Border, PatternFill
from openpyxl.cell.cell import MergedCell
import pandas as pd
from copy import copy  # for copying cell styles

# ------------------------------------------------------------
# Logging Configuration
# ------------------------------------------------------------
logging.basicConfig(
    level=logging.DEBUG,  # Change to INFO or ERROR for less verbosity
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)

# ------------------------------------------------------------
# Constants / Paths
# ------------------------------------------------------------
TEMPLATE_PATH = r"template_daily_recap.xlsx"  # Your template file
TEMPLATE_SHEET_NAME = "Template"
TOTAL_SHEET_NAME = "Total"
OUTPUT_DIR = "TrackedWorkLog"  # Folder for the generated workbook

# The row number that contains the desired data cell style (adjust if needed)
TEMPLATE_STYLE_ROW = 7

# ------------------------------------------------------------
# Safe Cell Writing (for merged cells)
# ------------------------------------------------------------
def safe_set_cell(ws, cell_ref, value):
    """
    Sets the value of a cell safely. If the cell is merged, only the top-left cell is updated.
    """
    cell = ws[cell_ref]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left.value = value
                return
    else:
        cell.value = value

# ------------------------------------------------------------
# Copy Cell Style from Source to Destination
# ------------------------------------------------------------
def copy_cell_style(src_cell, dst_cell):
    """
    Copies style from src_cell to dst_cell (includes font, fill, border, alignment, number format, etc.).
    """
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)

# ------------------------------------------------------------
# Clear Data from Worksheet (from row 7 downward)
# ------------------------------------------------------------
def clear_sheet_data(ws, start_row=7):
    """
    Clears cell values from start_row to the end of the sheet,
    preserving rows 1–6 (the template design and formatting).
    """
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            safe_set_cell(ws, cell.coordinate, None)

# ------------------------------------------------------------
# Apply Consistent Layout in Rows 1–6
# ------------------------------------------------------------
def apply_consistent_layout(ws, date_label):
    """
    Sets the top layout (rows 1–6) as follows:
      Row 1: A1 = "DATE", B1 = today's date (mm-dd-yyyy)
      Row 2: A2 = "Daily Recap ( - LA Office )"
      Row 3: A3 = "Date", B3 = date_label (user input)
      Row 4: A4 = "Name"
      Row 5: A5 = "Department"
      Row 6: The headers:
           "Number | Daily Work Description | Hr | Min | Complete | Follow up | Supervisor Comments"
    Freezes the pane at row 7.
    """
    today_str = datetime.date.today().strftime("%m-%d-%Y")
    safe_set_cell(ws, "A1", "DATE")
    safe_set_cell(ws, "B1", today_str)
    safe_set_cell(ws, "A2", "Daily Recap ( - LA Office )")
    safe_set_cell(ws, "A3", "Date")
    safe_set_cell(ws, "B3", date_label)
    safe_set_cell(ws, "A4", "Name")
    safe_set_cell(ws, "A5", "Department")
    headers = ["Number", "Daily Work Description", "Hr", "Min", "Complete", "Follow up", "Supervisor Comments"]
    for col_idx, header in enumerate(headers, start=1):
        cell_ref = ws.cell(row=6, column=col_idx).coordinate
        safe_set_cell(ws, cell_ref, header)
    ws.freeze_panes = "A7"

# ------------------------------------------------------------
# Check File Format for Required Columns
# ------------------------------------------------------------
def check_file_format(file_path):
    """
    Searches for a line that contains the required header:
      Number, Daily Work Description, Hr, Min, Complete, Follow up, Supervisor Comments.
    Exits if not found.
    """
    required_columns = [
        "Number", "Daily Work Description", "Hr", "Min", "Complete", "Follow up", "Supervisor Comments"
    ]
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
    except Exception as e:
        logging.error(f"Error reading file '{file_path}': {e}")
        exit(1)
    header_found = False
    for line in lines:
        stripped_line = line.strip()
        if all(col in stripped_line for col in required_columns):
            header_found = True
            break
    if not header_found:
        logging.error(f"File '{file_path}' does not contain the required header.")
        exit(1)
    logging.info(f"File '{file_path}' passed the format check.")

# ------------------------------------------------------------
# Prompt for Date Range (mm-dd-yyyy)
# ------------------------------------------------------------
def prompt_date_range():
    print("Enter the date range for your daily sheets (mm-dd-yyyy). Press Enter to use today's date.")
    start_str = input("  Start date (mm-dd-yyyy) [Optional]: ").strip()
    end_str = input("  End date   (mm-dd-yyyy) [Optional]: ").strip()
    if not start_str and not end_str:
        logging.info("No dates provided; using today's date for a single sheet.")
        return None, None
    if start_str and not end_str:
        try:
            sd = datetime.datetime.strptime(start_str, "%m-%d-%Y").date()
        except ValueError:
            logging.error("Invalid start date format. Use mm-dd-yyyy.")
            exit(1)
        return sd, None
    try:
        sd = datetime.datetime.strptime(start_str, "%m-%d-%Y").date()
        ed = datetime.datetime.strptime(end_str, "%m-%d-%Y").date()
    except ValueError:
        logging.error("Invalid date format. Use mm-dd-yyyy.")
        exit(1)
    if sd > ed:
        logging.error("Start date must not be later than end date.")
        exit(1)
    logging.info(f"Date range provided: {sd} to {ed}")
    return sd, ed

# ------------------------------------------------------------
# Prompt for File Paths (CSV/TXT)
# ------------------------------------------------------------
def prompt_file_paths():
    paths_str = input("Enter the path(s) to your CSV/TXT file(s) (comma-separated) [Optional]: ").strip()
    if not paths_str:
        logging.info("No data files provided. Daily sheets will be created empty.")
        return []
    file_paths = [fp.strip().strip('"') for fp in paths_str.split(",")]
    for fp in file_paths:
        if not os.path.exists(fp):
            logging.error(f"File not found: {fp}")
            exit(1)
    return file_paths

# ------------------------------------------------------------
# Prompt for Hourly Rate
# ------------------------------------------------------------
def prompt_rate():
    rate_str = input("Enter the hourly rate: ").strip()
    try:
        rate = float(rate_str)
    except ValueError:
        logging.error("Invalid rate. Please enter a numeric value.")
        exit(1)
    logging.info(f"Hourly rate: {rate}")
    return rate

# ------------------------------------------------------------
# Read CSV/TXT Data into a DataFrame
# ------------------------------------------------------------
def read_csv_data(data_file):
    """
    Checks the file format and reads a CSV or tab-delimited TXT file into a DataFrame.
    """
    check_file_format(data_file)
    try:
        with open(data_file, "r", encoding="utf-8") as f:
            lines = f.readlines()
        required_columns = [
            "Number", "Daily Work Description", "Hr", "Min", "Complete", "Follow up", "Supervisor Comments"
        ]
        header_index = None
        for i, line in enumerate(lines):
            if line.strip() and all(col in line for col in required_columns):
                header_index = i
                break
        if header_index is None:
            header_index = 0
        if data_file.lower().endswith(".txt"):
            df = pd.read_csv(data_file, sep='\t', skiprows=header_index, encoding="utf-8")
        else:
            df = pd.read_csv(data_file, skiprows=header_index, encoding="utf-8")
        logging.info(f"Data file '{data_file}' read with {len(df)} rows (skipped {header_index} rows).")
    except Exception as e:
        logging.error(f"Error reading data file '{data_file}': {e}")
        exit(1)
    return df

# ------------------------------------------------------------
# Combine Multiple CSV/TXT Files
# ------------------------------------------------------------
def combine_csv_data(file_paths):
    if not file_paths:
        return pd.DataFrame()
    dfs = []
    for fp in file_paths:
        df = read_csv_data(fp)
        dfs.append(df)
    combined_df = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
    if "Date" in combined_df.columns:
        combined_df["Date"] = pd.to_datetime(combined_df["Date"], errors="coerce").dt.date
        if combined_df["Date"].isna().sum() > 0:
            logging.warning("Some dates in the 'Date' column could not be parsed.")
    return combined_df

# ------------------------------------------------------------
# Create a List of Dates from Start to End
# ------------------------------------------------------------
def create_date_list(start_date, end_date):
    dates = []
    current = start_date
    while current <= end_date:
        dates.append(current)
        current += datetime.timedelta(days=1)
    return dates

# ------------------------------------------------------------
# Fill a Daily Sheet with Data (Copying Style from TEMPLATE_STYLE_ROW)
# ------------------------------------------------------------
def fill_daily_sheet(sheet, date_obj, data_df, start_row=7):
    """
    Populates a daily sheet:
      - Filters data (if a "Date" column exists) for date_obj.
      - Writes data starting at row 'start_row' (columns A–G).
      - For each written cell, copies the style from the cell in TEMPLATE_STYLE_ROW of the same column.
    Returns the last row used.
    """
    if not data_df.empty and "Date" in data_df.columns:
        day_df = data_df[data_df["Date"] == date_obj]
    else:
        day_df = data_df
    records = day_df.to_dict(orient="records")
    current = start_row
    for rec in records:
        for col_idx, key in enumerate(["Number", "Daily Work Description", "Hr", "Min", "Complete", "Follow up", "Supervisor Comments"], start=1):
            dest_cell = sheet.cell(row=current, column=col_idx)
            dest_cell.value = rec.get(key, "")
            # If the source (template) cell exists in TEMPLATE_STYLE_ROW, copy its style
            src_cell = sheet.cell(row=TEMPLATE_STYLE_ROW, column=col_idx)
            copy_cell_style(src_cell, dest_cell)
            # For the "Complete" column, adjust font color if needed.
            if key == "Complete":
                val = rec.get(key, "")
                if isinstance(val, str):
                    if val.strip().lower() == "yes":
                        dest_cell.font = Font(color="008000")
                    elif val.strip().lower() == "no":
                        dest_cell.font = Font(color="FF0000")
        current += 1
    last_row = current - 1
    logging.info(f"{date_obj}: Data written from row {start_row} to {last_row}.")
    return last_row

# ------------------------------------------------------------
# Update the "Total" Sheet with Daily Summaries
# ------------------------------------------------------------
def update_total_sheet(total_sheet, daily_info, rate):
    """
    Fills the 'Total' sheet with summary info:
      - Sets headers in cells B3:E3 (Date, Hour, Rate, Total Cost)
      - From row 4 onward, each date gets a row with formulas summing hours and minutes.
      - The Hour is referenced from each daily sheet's C4.
      - Total Cost is computed as Hour * rate.
    """
    # Clear cells from row 4 downward
    for row in total_sheet.iter_rows(min_row=4, max_row=total_sheet.max_row):
        for cell in row:
            cell.value = None

    row_idx = 4
    for sheet_name, (start_row, last_row) in sorted(daily_info.items()):
        if last_row < start_row:
            continue

        hour_formula = (
            f"=SUM('{sheet_name}'!C{start_row}:C{last_row}) + "
            f"(SUM('{sheet_name}'!D{start_row}:D{last_row})/60)"
        )
        safe_set_cell(total_sheet, f"B{row_idx}", sheet_name)
        safe_set_cell(total_sheet, f"C{row_idx}", hour_formula)
        safe_set_cell(total_sheet, f"D{row_idx}", rate)
        safe_set_cell(total_sheet, f"E{row_idx}", f"=C{row_idx}*D{row_idx}")
        row_idx += 1

# ------------------------------------------------------------
# Create or Update the "Total" Sheet
# ------------------------------------------------------------
def create_or_update_total_sheet(wb, daily_info, rate):
    """
    Ensures a 'Total' sheet exists, writes summary headers, populates rows,
    and freezes rows above row 4.
    """
    if TOTAL_SHEET_NAME in wb.sheetnames:
        total_sheet = wb[TOTAL_SHEET_NAME]
    else:
        total_sheet = wb.create_sheet(TOTAL_SHEET_NAME)

    safe_set_cell(total_sheet, "B3", "Date")
    safe_set_cell(total_sheet, "C3", "Hour")
    safe_set_cell(total_sheet, "D3", "Rate")
    safe_set_cell(total_sheet, "E3", "Total Cost")
    update_total_sheet(total_sheet, daily_info, rate)
    total_sheet.freeze_panes = total_sheet["A4"]
    return total_sheet

# ------------------------------------------------------------
# Main Workflow
# ------------------------------------------------------------
def main():
    """
    Main workflow:
      1) Prompt for date range (or use today's date if none provided)
      2) Prompt for file paths (CSV/TXT; optional)
      3) Prompt for hourly rate
      4) Load the template workbook
      5) Combine file data into a DataFrame (if any)
      6) Create daily sheets based on provided dates:
           - Apply the consistent top layout (rows 1–6)
           - Clear any data from row 7 downward
           - Write data (filtered by date if available) starting at row 7
           - Insert a formula in cell C4: =SUM(C7:C_last) + (SUM(D7:D_last)/60)
      7) Create/Update the single "Total" sheet summarizing each daily sheet
      8) Hide the template sheet and save the workbook in OUTPUT_DIR
    """
    start_date, end_date = prompt_date_range()
    file_paths = prompt_file_paths()
    rate = prompt_rate()

    if not os.path.exists(TEMPLATE_PATH):
        logging.error(f"Template file not found: {TEMPLATE_PATH}")
        exit(1)
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    if TEMPLATE_SHEET_NAME not in wb.sheetnames:
        logging.error(f"No sheet named '{TEMPLATE_SHEET_NAME}' in the template.")
        exit(1)

    combined_df = combine_csv_data(file_paths)
    daily_info = {}

    if start_date is None and end_date is None:
        single_date = datetime.date.today()
        date_label = single_date.strftime("%m-%d-%Y")
        date_list = [single_date]
    elif start_date is not None and end_date is None:
        date_label = start_date.strftime("%m-%d-%Y")
        date_list = [start_date]
    else:
        date_list = create_date_list(start_date, end_date)
        date_label = f"{start_date.strftime('%m-%d-%Y')}_to_{end_date.strftime('%m-%d-%Y')}"

    for day in date_list:
        sheet_name = day.strftime("%m-%d-%Y")
        new_sheet = wb.copy_worksheet(wb[TEMPLATE_SHEET_NAME])
        new_sheet.title = sheet_name
        apply_consistent_layout(new_sheet, date_label)
        clear_sheet_data(new_sheet, start_row=7)
        last_row = fill_daily_sheet(new_sheet, date_obj=day, data_df=combined_df, start_row=7)
        daily_info[sheet_name] = (7, last_row)
        if last_row >= 7:
            new_sheet["C4"] = f"=SUM(C7:C{last_row}) + (SUM(D7:D{last_row})/60)"
        else:
            new_sheet["C4"] = 0

    create_or_update_total_sheet(wb, daily_info, rate)
    wb[TEMPLATE_SHEET_NAME].sheet_state = "hidden"

    if len(date_list) == 1:
        output_filename = f"{date_list[0].strftime('%m-%d-%Y')}.xlsx"
    else:
        output_filename = f"{date_list[0].strftime('%m-%d-%Y')}_to_{date_list[-1].strftime('%m-%d-%Y')}.xlsx"
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    output_filepath = os.path.join(OUTPUT_DIR, output_filename)
    try:
        wb.save(output_filepath)
    except PermissionError as e:
        logging.error(f"Permission error saving '{output_filepath}': {e}")
        exit(1)
    logging.info(f"Workbook '{output_filepath}' created successfully.")

# ------------------------------------------------------------
# Entry Point
# ------------------------------------------------------------
if __name__ == "__main__":
    main()
