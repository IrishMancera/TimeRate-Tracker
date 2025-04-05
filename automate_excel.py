import os
import re
import datetime
import logging
import openpyxl
from openpyxl.styles import Font
from openpyxl.cell.cell import MergedCell
import pandas as pd

# ------------------------------------------------------------
# Logging Configuration
# ------------------------------------------------------------
logging.basicConfig(
    level=logging.DEBUG,  # Change to INFO or ERROR for less verbosity
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)

# ------------------------------------------------------------
# Constants / Paths
# ------------------------------------------------------------
TEMPLATE_PATH = r"template_daily_recap.xlsx"  # Adjust if stored elsewhere
TEMPLATE_SHEET_NAME = "Template"
TOTAL_SHEET_NAME = "Total"
OUTPUT_DIR = "TrackedWorkLog"  # All generated files will be saved here

# ------------------------------------------------------------
# Safe Cell Writing (for merged cells)
# ------------------------------------------------------------
def safe_set_cell(ws, cell_ref, value):
    """
    Safely sets the value of a cell. If the cell is merged, update the top-left cell.
    """
    cell = ws[cell_ref]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left.value = value
                return
    else:
        cell.value = value

# ------------------------------------------------------------
# Clear Data from Copied Worksheet
# ------------------------------------------------------------
def clear_sheet_data(sheet, start_row=7):
    """
    Clears cell values from start_row to the end of the sheet.
    This prevents copying any sample data from the TEMPLATE.
    """
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row):
        for cell in row:
            cell.value = None

# ------------------------------------------------------------
# Check File Format (Modified)
# ------------------------------------------------------------
def check_file_format(file_path):
    """
    Instead of requiring the first 7 lines to match a fixed snippet,
    this function now searches for any line that contains the required header items:
    
      - Number
      - Daily Work Description
      - Hr
      - Min
      - Complete
      - Follow up
      - Supervisor Comments

    If such a line is found, the file is accepted. Otherwise, the process terminates.
    """
    required_columns = [
        "Number",
        "Daily Work Description",
        "Hr",
        "Min",
        "Complete",
        "Follow up",
        "Supervisor Comments"
    ]

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
    except Exception as e:
        logging.error(f"Error reading file '{file_path}': {e}")
        exit(1)

    header_found = False
    for line in lines:
        stripped_line = line.strip()
        if all(col in stripped_line for col in required_columns):
            header_found = True
            break

    if not header_found:
        logging.error(
            f"File '{file_path}' does not contain the required header: "
            f"'Number  Daily Work Description  Hr      Min     Complete        Follow up       Supervisor Comments'"
        )
        exit(1)

    logging.info(f"File '{file_path}' passed the format check.")

# ------------------------------------------------------------
# Prompt for Date Range (mm-dd-yyyy) [Modified to allow empty inputs]
# ------------------------------------------------------------
def prompt_date_range():
    """
    Prompts for a start and end date in mm-dd-yyyy format, allowing for empty input.
    Returns (start_date, end_date) as either date objects or None.
    
    Possible outcomes:
      - Both provided -> returns (start_date, end_date)
      - Only start date -> returns (start_date, None)
      - Neither -> returns (None, None)
    """
    print("Enter the date range for your daily sheets (mm-dd-yyyy). Press Enter to skip.")
    start_date_str = input("  Start date (mm-dd-yyyy) [Optional]: ").strip()
    end_date_str   = input("  End date   (mm-dd-yyyy) [Optional]: ").strip()

    if not start_date_str and not end_date_str:
        logging.info("No start/end date provided; will use today's date for a single sheet.")
        return None, None

    if start_date_str and not end_date_str:
        try:
            start_date = datetime.datetime.strptime(start_date_str, "%m-%d-%Y").date()
        except ValueError:
            logging.error("Invalid start date format. Please use mm-dd-yyyy.")
            exit(1)
        logging.info(f"Only start date provided: {start_date}")
        return start_date, None

    try:
        start_date = datetime.datetime.strptime(start_date_str, "%m-%d-%Y").date()
        end_date   = datetime.datetime.strptime(end_date_str, "%m-%d-%Y").date()
    except ValueError:
        logging.error("Invalid date format. Please use mm-dd-yyyy.")
        exit(1)

    if start_date > end_date:
        logging.error("Start date must not be later than end date.")
        exit(1)

    return start_date, end_date

# ------------------------------------------------------------
# Prompt for File Paths (CSV/TXT) [Modified to allow empty input]
# ------------------------------------------------------------
def prompt_file_paths():
    """
    Prompts for one or more CSV/TXT file paths (comma-separated) and verifies each exists.
    Returns a list of file paths (possibly empty if user skipped).
    """
    file_paths_str = input("Enter the path(s) to the data file(s) (CSV/TXT, comma-separated) [Optional]: ").strip()
    if not file_paths_str:
        logging.info("No data files provided. Will create empty daily sheets.")
        return []

    file_paths = [fp.strip().strip('"') for fp in file_paths_str.split(",")]
    for fp in file_paths:
        if not os.path.exists(fp):
            logging.error(f"File not found: {fp}")
            exit(1)
    return file_paths

# ------------------------------------------------------------
# Prompt for Hourly Rate
# ------------------------------------------------------------
def prompt_rate():
    """
    Prompts for an hourly rate and returns it as float.
    """
    rate_str = input("Enter the hourly rate: ").strip()
    try:
        rate = float(rate_str)
    except ValueError:
        logging.error("Invalid rate. Please enter a numeric value.")
        exit(1)
    logging.info(f"Hourly rate: {rate}")
    return rate

# ------------------------------------------------------------
# Read CSV/TXT into DataFrame (Refined)
# ------------------------------------------------------------
def read_csv_data(data_file):
    """
    Reads a CSV or TXT file into a Pandas DataFrame.
    First, it checks that the file contains the required header.
    Then, it scans the file to determine how many rows to skip so that the header row becomes the first row.
    If the file ends with '.txt', it assumes tab-delimited; otherwise, comma-delimited.
    """
    # Check file format first
    check_file_format(data_file)

    try:
        with open(data_file, "r", encoding="utf-8") as f:
            lines = f.readlines()
        # Find the index of the first line that contains all required columns.
        required_columns = [
            "Number",
            "Daily Work Description",
            "Hr",
            "Min",
            "Complete",
            "Follow up",
            "Supervisor Comments"
        ]
        header_line_index = None
        for i, line in enumerate(lines):
            if line.strip() and all(col in line for col in required_columns):
                header_line_index = i
                break
        if header_line_index is None:
            header_line_index = 0  # fallback if header is not found (shouldn't happen)
        # Read the CSV/TXT using the detected header line as the first row.
        if data_file.lower().endswith(".txt"):
            df = pd.read_csv(data_file, sep='\t', skiprows=header_line_index)
        else:
            df = pd.read_csv(data_file, skiprows=header_line_index)
        logging.info(f"Data file '{data_file}' read with {len(df)} rows (skipped {header_line_index} rows).")
    except Exception as e:
        logging.error(f"Error reading data file '{data_file}': {e}")
        exit(1)
    return df

# ------------------------------------------------------------
# Combine multiple CSV/TXT files
# ------------------------------------------------------------
def combine_csv_data(file_paths):
    """
    Reads and concatenates multiple CSV/TXT files into a single DataFrame.
    If no files are provided, returns an empty DataFrame.
    """
    if not file_paths:
        return pd.DataFrame()  # Empty

    df_list = []
    for fp in file_paths:
        df = read_csv_data(fp)
        df_list.append(df)
    if df_list:
        combined_df = pd.concat(df_list, ignore_index=True)
    else:
        combined_df = pd.DataFrame()
    return combined_df

# ------------------------------------------------------------
# Create a list of date objects from start_date to end_date
# ------------------------------------------------------------
def create_date_list(start_date, end_date):
    """
    Creates a list of date objects from start_date to end_date (inclusive).
    """
    date_list = []
    current = start_date
    while current <= end_date:
        date_list.append(current)
        current += datetime.timedelta(days=1)
    return date_list

# ------------------------------------------------------------
# Filter DataFrame by the given date range (if 'Date' column exists)
# ------------------------------------------------------------
def filter_df_by_date(df, start_date, end_date):
    """
    If a 'Date' column exists, parses it as date and filters the DataFrame by the given date range.
    Returns (filtered_df, has_date_column).
    If df is empty or no date column, returns original df with has_date_column=False.
    """
    if df.empty:
        return df, False

    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.date
        original_count = len(df)
        df = df[(df['Date'] >= start_date) & (df['Date'] <= end_date)]
        logging.info(f"Filtered from {original_count} rows to {len(df)} rows by 'Date' column.")
        return df, True
    else:
        logging.info("No 'Date' column found; applying all data to each date in the range.")
        return df, False

# ------------------------------------------------------------
# Fill a daily sheet with data
# ------------------------------------------------------------
def fill_daily_sheet(sheet, date_obj, data_rows, is_dataframe, start_row=7, fallback_date=None):
    """
    Populates a daily sheet:
      - Sets cell B1 to the date (mm-dd-yyyy) if data exists, else today's date.
      - If no data for that date, also writes fallback_date to B3.
      - Writes typical headers in row 6, then data from row 7 onward.
    Returns the last row used.
    """
    if is_dataframe and not data_rows.empty:
        if 'Date' in data_rows.columns:
            day_df = data_rows[data_rows['Date'] == date_obj]
        else:
            day_df = data_rows
        records = day_df.to_dict(orient='records')
    else:
        records = []

    if not records:
        today_str = datetime.date.today().strftime("%m-%d-%Y")
        sheet["B1"] = today_str
        if fallback_date is not None:
            sheet["B3"] = fallback_date.strftime("%m-%d-%Y")
    else:
        sheet["B1"] = date_obj.strftime("%m-%d-%Y")

    sheet.freeze_panes = sheet["A7"]
    headers = ["Number", "Daily Work Description", "Hr", "Min", "Complete", "Follow up", "Supervisor Comments"]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=6, column=col_idx).value = header

    current_row = start_row
    for row_dict in records:
        sheet.cell(row=current_row, column=1).value = row_dict.get("Number", "")
        sheet.cell(row=current_row, column=2).value = row_dict.get("Daily Work Description", "")
        sheet.cell(row=current_row, column=3).value = row_dict.get("Hr", "")
        sheet.cell(row=current_row, column=4).value = row_dict.get("Min", "")

        complete_val = row_dict.get("Complete", "")
        complete_cell = sheet.cell(row=current_row, column=5)
        complete_cell.value = complete_val
        if isinstance(complete_val, str):
            if complete_val.lower() == "yes":
                complete_cell.font = Font(color="008000")  # green
            elif complete_val.lower() == "no":
                complete_cell.font = Font(color="FF0000")  # red

        sheet.cell(row=current_row, column=6).value = row_dict.get("Follow up", "")
        sheet.cell(row=current_row, column=7).value = row_dict.get("Supervisor Comments", "")
        current_row += 1

    last_row = current_row - 1
    logging.info(f"{date_obj}: Populated rows {start_row} to {last_row}.")
    return last_row

# ------------------------------------------------------------
# Update the "Total" sheet with daily info
# ------------------------------------------------------------
def update_total_sheet(total_sheet, daily_info, rate):
    """
    Fills the 'Total' sheet with summary info:
      - Sets headers in cells B3:E3 (Date, Hour, Rate, Total Cost)
      - From row 4 onward, each date gets a row with formulas summing hours and minutes.
    """
    for row in total_sheet.iter_rows(min_row=4, max_row=total_sheet.max_row):
        for cell in row:
            cell.value = None

    row_idx = 4
    for sheet_name, (start_row, last_row) in sorted(daily_info.items()):
        if last_row < start_row:
            continue

        hour_formula = (
            f"=SUM('{sheet_name}'!C{start_row}:C{last_row}) + "
            f"(SUM('{sheet_name}'!D{start_row}:D{last_row})/60)"
        )
        safe_set_cell(total_sheet, f"B{row_idx}", sheet_name)
        safe_set_cell(total_sheet, f"C{row_idx}", hour_formula)
        safe_set_cell(total_sheet, f"D{row_idx}", rate)
        safe_set_cell(total_sheet, f"E{row_idx}", f"=C{row_idx}*D{row_idx}")
        row_idx += 1

# ------------------------------------------------------------
# Create or update the "Total" sheet
# ------------------------------------------------------------
def create_or_update_total_sheet(wb, daily_info, rate):
    """
    Ensures a 'Total' sheet exists, writes summary headers, populates rows,
    and freezes rows above row 4.
    """
    if TOTAL_SHEET_NAME in wb.sheetnames:
        total_sheet = wb[TOTAL_SHEET_NAME]
    else:
        total_sheet = wb.create_sheet(TOTAL_SHEET_NAME)

    safe_set_cell(total_sheet, "B3", "Date")
    safe_set_cell(total_sheet, "C3", "Hour")
    safe_set_cell(total_sheet, "D3", "Rate")
    safe_set_cell(total_sheet, "E3", "Total Cost")
    update_total_sheet(total_sheet, daily_info, rate)
    total_sheet.freeze_panes = total_sheet["A4"]
    return total_sheet

# ------------------------------------------------------------
# Main Workflow
# ------------------------------------------------------------
def main():
    """
    Main workflow:
      1) Prompt for date range (or use today's date if none provided)
      2) Prompt for file paths (CSV/TXT; optional)
      3) Prompt for hourly rate
      4) Load the template workbook
      5) Combine file data into a DataFrame (if any)
      6) Create daily sheets based on provided dates
      7) Create/Update the "Total" sheet
      8) Hide the template sheet
      9) Save the workbook inside the "TrackedWorkLog" folder
    """
    start_date, end_date = prompt_date_range()
    file_paths = prompt_file_paths()
    rate = prompt_rate()

    if not os.path.exists(TEMPLATE_PATH):
        logging.error(f"Template file not found: {TEMPLATE_PATH}")
        exit(1)
    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    if TEMPLATE_SHEET_NAME not in wb.sheetnames:
        logging.error(f"No sheet named '{TEMPLATE_SHEET_NAME}' in {TEMPLATE_PATH}")
        exit(1)

    combined_df = combine_csv_data(file_paths)
    daily_info = {}

    # Scenario: No dates provided – use today's date
    if start_date is None and end_date is None:
        single_date = datetime.date.today()
        sheet_name = single_date.strftime("%m-%d-%Y")
        new_sheet = wb.copy_worksheet(wb[TEMPLATE_SHEET_NAME])
        new_sheet.title = sheet_name
        clear_sheet_data(new_sheet, start_row=7)
        last_row = fill_daily_sheet(new_sheet, date_obj=single_date, data_rows=combined_df,
                                    is_dataframe=True, start_row=7, fallback_date=single_date)
        daily_info[sheet_name] = (7, last_row)
        create_or_update_total_sheet(wb, daily_info, rate)
        wb[TEMPLATE_SHEET_NAME].sheet_state = "hidden"
        output_filename = f"{single_date.strftime('%m-%d-%Y')}.xlsx"
        if not os.path.exists(OUTPUT_DIR):
            os.makedirs(OUTPUT_DIR)
        output_filepath = os.path.join(OUTPUT_DIR, output_filename)
        try:
            wb.save(output_filepath)
        except PermissionError as e:
            logging.error(f"Permission error saving '{output_filepath}': {e}")
            exit(1)
        logging.info(f"Workbook '{output_filepath}' created successfully.")
        return

    # Scenario: Only start_date provided (single day)
    if start_date is not None and end_date is None:
        single_date = start_date
        if not combined_df.empty and 'Date' in combined_df.columns:
            combined_df['Date'] = pd.to_datetime(combined_df['Date'], errors='coerce').dt.date
            combined_df = combined_df[combined_df['Date'] == single_date]
        sheet_name = single_date.strftime("%m-%d-%Y")
        new_sheet = wb.copy_worksheet(wb[TEMPLATE_SHEET_NAME])
        new_sheet.title = sheet_name
        clear_sheet_data(new_sheet, start_row=7)
        last_row = fill_daily_sheet(new_sheet, date_obj=single_date, data_rows=combined_df,
                                    is_dataframe=True, start_row=7, fallback_date=single_date)
        daily_info[sheet_name] = (7, last_row)
        create_or_update_total_sheet(wb, daily_info, rate)
        wb[TEMPLATE_SHEET_NAME].sheet_state = "hidden"
        output_filename = f"{single_date.strftime('%m-%d-%Y')}.xlsx"
        if not os.path.exists(OUTPUT_DIR):
            os.makedirs(OUTPUT_DIR)
        output_filepath = os.path.join(OUTPUT_DIR, output_filename)
        try:
            wb.save(output_filepath)
        except PermissionError as e:
            logging.error(f"Permission error saving '{output_filepath}': {e}")
            exit(1)
        logging.info(f"Workbook '{output_filepath}' created successfully.")
        return

    # Scenario: Both start_date and end_date provided (date range)
    if not combined_df.empty and 'Date' in combined_df.columns:
        combined_df['Date'] = pd.to_datetime(combined_df['Date'], errors='coerce').dt.date
        original_count = len(combined_df)
        combined_df = combined_df[(combined_df['Date'] >= start_date) & (combined_df['Date'] <= end_date)]
        logging.info(f"Filtered from {original_count} rows to {len(combined_df)} rows by date range.")

    date_list = create_date_list(start_date, end_date)
    for day in date_list:
        sheet_name = day.strftime("%m-%d-%Y")
        new_sheet = wb.copy_worksheet(wb[TEMPLATE_SHEET_NAME])
        new_sheet.title = sheet_name
        clear_sheet_data(new_sheet, start_row=7)
        fallback = day if (start_date == end_date) else None
        last_row = fill_daily_sheet(new_sheet, date_obj=day, data_rows=combined_df,
                                    is_dataframe=True, start_row=7, fallback_date=fallback)
        daily_info[sheet_name] = (7, last_row)

    create_or_update_total_sheet(wb, daily_info, rate)
    wb[TEMPLATE_SHEET_NAME].sheet_state = "hidden"

    if start_date == end_date:
        output_filename = f"{start_date.strftime('%m-%d-%Y')}.xlsx"
    else:
        output_filename = f"{start_date.strftime('%m-%d-%Y')}_to_{end_date.strftime('%m-%d-%Y')}.xlsx"
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    output_filepath = os.path.join(OUTPUT_DIR, output_filename)
    try:
        wb.save(output_filepath)
    except PermissionError as e:
        logging.error(f"Permission error saving '{output_filepath}': {e}")
        exit(1)
    logging.info(f"Workbook '{output_filepath}' created successfully.")

# ------------------------------------------------------------
# Entry Point
# ------------------------------------------------------------
if __name__ == "__main__":
    main()
    